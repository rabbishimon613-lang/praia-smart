"""Scrape Brazilian beach water-quality (balneabilidade) reports.

Sources wired:
  - IMA-SC  (Santa Catarina) → floripa-barra-lagoa, balneario-camboriu,
            balneario-rincao. Weekly Fri PDF.
  - CETESB  (São Paulo)      → santos-gonzaga, guaruja-enseada,
            praia-grande-boqueirao. Weekly HTML map page.
  - INEA    (Rio de Janeiro) → copa-p5, leblon-mirante, sao-conrado,
            cabo-frio-forte. Weekly PDF linked from landing page.
  - INEMA   (Bahia)          → morro-terceira, salvador-itapua,
            salvador-stella-maris, porto-seguro-taperapua. INEMA's PDF
            URLs aren't reliably discoverable, but the agency publishes
            a weekly *news article* on www.ba.gov.br/inema with the
            full bulletin in prose form — we parse that.

  - FEPAM   (Rio Grande do Sul) → tramandai, capao-da-canoa. Seasonal:
            weekly Friday bulletin only during Operação Verão Total
            (Dec–Feb). Off-season we return sem_dados stubs with
            off_season=True.

natal-ponta-negra (IDEMA-RN) gets a sem_dados placeholder for now —
no clean URL pattern.

INEMA "Vai dar Praia" app: probed api.inema.ba.gov.br (NXDOMAIN),
servicos.inema.ba.gov.br (timeout), vaidarpraia.inema.ba.gov.br
(NXDOMAIN), and the Brasil.IO balneabilidade-bahia dataset (auth-only).
No public JSON backend found in May 2026 — sticking with the news-article
scraper.

SEMACE Digital app: probed api.semace.ce.gov.br (timeout),
mobile.semace.ce.gov.br (503), servicos.semace.ce.gov.br (timeout).
No public JSON backend found — sticking with the PDF scraper.

Stdlib + curl + pdftotext (poppler) only, to match other workers/ scripts.

Run:  python3 workers/balneabilidade.py
Out:  data/balneabilidade.json
"""
import json
import os
import re
import subprocess
import sys
import time
from datetime import date, datetime, timedelta
from html.parser import HTMLParser

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "balneabilidade.json")

UA = "praiasmart/0.1 (+balneabilidade)"
TIMEOUT = 30

# All beaches we care about — anything not filled in by a scraper
# gets a sem_dados stub at the end of build().
ALL_BEACH_IDS = [
    "copa-p5", "leblon-mirante", "sao-conrado", "cabo-frio-forte",
    "ipanema-posto-9", "barra-tijuca-posto-3", "buzios-geriba",
    "balneario-camboriu", "balneario-rincao", "floripa-barra-lagoa",
    "jurere-internacional",
    "morro-terceira", "santos-gonzaga", "guaruja-enseada",
    "praia-grande-boqueirao", "natal-ponta-negra",
    "salvador-itapua", "salvador-stella-maris", "porto-seguro-taperapua",
    "recife-boa-viagem", "ipojuca-porto-galinhas",
    "fortaleza-iracema-meireles",
    "maceio-pajucara",
    "tramandai", "capao-da-canoa",
]

# Tier-2 sources we don't scrape yet: still report which agency would own them.
DEFAULT_SOURCE = {
    "morro-terceira": "INEMA",
    "salvador-itapua": "INEMA",
    "salvador-stella-maris": "INEMA",
    "porto-seguro-taperapua": "INEMA",
    "natal-ponta-negra": "IDEMA",
    "ipanema-posto-9": "INEA",
    "barra-tijuca-posto-3": "INEA",
    "buzios-geriba": "INEA",
    "jurere-internacional": "IMA-SC",
    "recife-boa-viagem": "CPRH",
    "ipojuca-porto-galinhas": "CPRH",
    "fortaleza-iracema-meireles": "SEMACE",
    "maceio-pajucara": "IMA-AL",
    "tramandai": "FEPAM",
    "capao-da-canoa": "FEPAM",
}


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------

def curl(url, binary=False):
    """Fetch a URL via curl. Returns text (decoded tolerantly) or bytes."""
    args = ["curl", "-sSfL", "-A", UA, "-m", str(TIMEOUT), url]
    raw = subprocess.run(args, capture_output=True, check=True,
                         timeout=TIMEOUT + 5).stdout
    if binary:
        return raw
    # Try utf-8 then latin-1 (CETESB and many .gov.br sites are cp1252/latin-1).
    for enc in ("utf-8", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def pdf_to_text(pdf_bytes):
    """Use poppler's pdftotext (layout-preserving) to extract text."""
    p = subprocess.run(
        ["pdftotext", "-layout", "-nopgbrk", "-", "-"],
        input=pdf_bytes, capture_output=True, timeout=30, check=True,
    )
    return p.stdout.decode("utf-8", errors="replace")


def classify(token):
    """Map agency status text → our normalized status."""
    if not token:
        return None
    t = token.lower()
    if "própr" in t or "propr" in t or t.strip() in {"p", "ok"}:
        # CETESB uses "Própria" / IMA uses "PRÓPRIA" / INEA same.
        return "propria"
    if "impr" in t or "imprópr" in t or t.strip() in {"i"}:
        return "impropria"
    if "alerta" in t or "atenção" in t or "atencao" in t:
        return "alerta"
    return None


# -----------------------------------------------------------------------------
# IMA-SC
# -----------------------------------------------------------------------------
# Weekly PDF: https://balneabilidade.ima.sc.gov.br/relatorio/downloadPDF/YYYY-MM-DD
# Try recent Fridays going back ~5 weeks.

IMA_NAME_MAP = {
    # substring match against PDF row text → beach_id
    "barra da lagoa": "floripa-barra-lagoa",
    "balneário camboriú": "balneario-camboriu",
    "balneario camboriu": "balneario-camboriu",
    "rincão": "balneario-rincao",
    "rincao": "balneario-rincao",
    "jurerê": "jurere-internacional",
    "jurere": "jurere-internacional",
}
IMA_PREFERRED_POINT = {
    # If multiple sample points hit, prefer one matching this hint
    "balneario-camboriu": "marambaia",  # matches Hotel Marambaia in postos.csv
    "floripa-barra-lagoa": None,
    "balneario-rincao": None,
    "jurere-internacional": "internacional",
}


def _recent_fridays(n=6):
    today = date.today()
    # weekday(): Mon=0 ... Fri=4
    offset = (today.weekday() - 4) % 7
    last_fri = today - timedelta(days=offset)
    return [last_fri - timedelta(days=7 * i) for i in range(n)]


def fetch_ima_sc():
    out = {}
    for d in _recent_fridays(6):
        url = f"https://balneabilidade.ima.sc.gov.br/relatorio/downloadPDF/{d.isoformat()}"
        try:
            pdf = curl(url, binary=True)
        except subprocess.CalledProcessError:
            continue
        if not pdf.startswith(b"%PDF"):
            continue
        try:
            text = pdf_to_text(pdf)
        except Exception as e:
            print(f"  IMA-SC: pdftotext failed for {d}: {e}", file=sys.stderr)
            continue
        report_date = d.isoformat()
        print(f"  IMA-SC: parsing {d}", file=sys.stderr)
        out = _parse_ima_text(text, url, report_date)
        if out:
            return out
    return out


def _parse_ima_text(text, url, report_date):
    """IMA PDFs list lines like:
        Florianópolis  Barra da Lagoa  Em frente ao posto 7  PRÓPRIA
    We do a best-effort substring match per beach id.
    """
    out = {}
    candidates = {bid: [] for bid in set(IMA_NAME_MAP.values())}
    for line in text.splitlines():
        low = line.lower()
        status = None
        if "própr" in low or "propr" in low:
            status = "propria" if "impr" not in low else "impropria"
        elif "impróp" in low or "improp" in low:
            status = "impropria"
        if not status:
            continue
        for needle, bid in IMA_NAME_MAP.items():
            if needle in low:
                candidates[bid].append((line.strip(), status))
                break
    for bid, hits in candidates.items():
        if not hits:
            continue
        chosen = hits[0]
        hint = IMA_PREFERRED_POINT.get(bid)
        if hint:
            for line, st in hits:
                if hint in line.lower():
                    chosen = (line, st)
                    break
        line, status = chosen
        out[bid] = {
            "status": status,
            "source": "IMA-SC",
            "report_date": report_date,
            "ecoli": None,
            "url": url,
            "confidence": "oficial",
        }
    return out


# -----------------------------------------------------------------------------
# CETESB
# -----------------------------------------------------------------------------

CETESB_BASE = "https://sistemasinter.cetesb.sp.gov.br/praias/mapa_praias/respraia1.asp"
CETESB_URL = CETESB_BASE  # for entry "url" field

# CETESB groups beaches per-city; we hit one URL per relevant city.
# city_code → (CETESB ?praia= code, list of (beach_name_substring, beach_id))
CETESB_CITIES = [
    ("ST", [
        ("gonzaga", "santos-gonzaga"),
    ]),
    ("GJ", [
        # In CETESB's Guarujá list, the Enseada points appear as
        # "ENSEADA-..." (Av. Atlântica matches the Enseada P11 area roughly).
        # First-hit on plain "enseada" is fine — they're contiguous.
        ("enseada", "guaruja-enseada"),
    ]),
    ("PG", [
        ("boqueir", "praia-grande-boqueirao"),
    ]),
]


_CETESB_ICON_STATUS = {
    "verde": "propria",
    "vermelha": "impropria",
    "vermelho": "impropria",
    "amarela": "alerta",
    "amarelo": "alerta",
}


def fetch_cetesb():
    # Pair each status-icon img with the *next* beach-name <td>.
    pattern = re.compile(
        r'src="(verde|vermelha|vermelho|amarela|amarelo)\.gif"'
        r'.*?<td[^>]*>(?:\s*<[^>]+>\s*)*\s*([A-Za-zÀ-ÿ0-9][^<]{1,60}?)\s*<',
        re.I | re.S,
    )
    out = {}
    report_date = None
    for code, beaches in CETESB_CITIES:
        url = f"{CETESB_BASE}?praia={code}"
        try:
            html = curl(url)
        except subprocess.CalledProcessError as e:
            print(f"  CETESB {code}: HTTP err {e}", file=sys.stderr)
            continue
        if report_date is None:
            m = re.search(r"(\d{2}/\d{2}/\d{4})", html)
            if m:
                try:
                    report_date = datetime.strptime(
                        m.group(1), "%d/%m/%Y").date().isoformat()
                except ValueError:
                    pass
        pairs = pattern.findall(html)
        for needle, bid in beaches:
            if bid in out:
                continue
            for icon, name in pairs:
                low = name.lower().strip()
                if needle in low:
                    status = _CETESB_ICON_STATUS.get(icon.lower())
                    if not status:
                        continue
                    out[bid] = {
                        "status": status,
                        "source": "CETESB",
                        "report_date": report_date,
                        "ecoli": None,
                        "url": url,
                        "confidence": "oficial",
                    }
                    break
    return out


# -----------------------------------------------------------------------------
# INEA
# -----------------------------------------------------------------------------

INEA_LANDING = "https://www.inea.rj.gov.br/ar-agua-e-solo/balneabilidade-das-praias/"

# CONAMA 274/2000 single-sample limit for Enterococos in saltwater:
# ≤100 NMP/100mL → própria; >100 → imprópria.
INEA_ENT_LIMIT = 100

# Station codes (Barra/Zona Sul "long" sheet) per target beach. We pick the
# WORST (highest enterococos) reading across the listed codes for the latest
# sampling date.
INEA_ZS_STATIONS = {
    "copa-p5": ["CP05"],
    "leblon-mirante": ["LM02"],
    "sao-conrado": ["PP10"],            # Pepino = São Conrado
    "ipanema-posto-9": ["IP03", "IP06", "IP10"],
    "barra-tijuca-posto-3": ["BT00", "BT01"],
}

# "Wide" sheets (col A = praia, col B = localização, col C = código, cols D… = dates).
# beach_id → (sheet_name, list of station_code prefixes to consider, worst-of)
INEA_WIDE = {
    "cabo-frio-forte": ("C.Frio (Enter.) desde 2011",
                        ["CF-02", "CF-03", "CF-04"]),  # Forte (3 pontos)
    "buzios-geriba":   ("Búzios (Enter.) desde 2011",
                        ["BZ-04"]),                    # Geribá
}

INEA_ZS_SHEET = "Barra e Z.Sul (Enterococos) "


def _ent_status(value):
    """Map Enterococos NMP/100mL → status using CONAMA 274/2000."""
    if value is None:
        return None
    try:
        n = int(value)
    except (TypeError, ValueError):
        # "NR" (não realizado) and other strings → no data
        return None
    return "propria" if n <= INEA_ENT_LIMIT else "impropria"


def _inea_find_xlsx_url(html):
    links = re.findall(r'href="([^"]+\.xlsx?[^"]*)"', html, re.I)
    if not links:
        return None
    # Prefer links mentioning "praias" / "monitoradas" / a recent year.
    def score(u):
        low = u.lower()
        s = 0
        if "praias" in low or "monitoradas" in low:
            s += 10
        if "2026" in low:
            s += 3
        elif "2025" in low:
            s += 2
        if low.endswith(".xlsx") or ".xlsx" in low:
            s += 1
        return s
    return sorted(set(links), key=score, reverse=True)[0]


def _inea_parse_zsul(wb, url):
    """Parse the long-format 'Barra e Z.Sul' sheet: dates in col A, station
    codes in row 2, values per cell."""
    if INEA_ZS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[INEA_ZS_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {}
    header = rows[1]  # row index 1 = station codes
    code_to_col = {}
    for i, v in enumerate(header):
        if isinstance(v, str) and v.strip():
            code_to_col[v.strip()] = i

    out = {}
    for bid, codes in INEA_ZS_STATIONS.items():
        cols = [code_to_col[c] for c in codes if c in code_to_col]
        if not cols:
            continue
        # Walk rows bottom-up, pick the most recent date where ≥1 code has a numeric value.
        for r in reversed(rows):
            if not hasattr(r[0], "year"):
                continue
            vals = []
            for c in cols:
                if c >= len(r):
                    continue
                v = r[c]
                if isinstance(v, (int, float)):
                    vals.append(int(v))
            if not vals:
                continue
            worst = max(vals)
            status = _ent_status(worst)
            if not status:
                continue
            out[bid] = {
                "status": status,
                "source": "INEA",
                "report_date": r[0].date().isoformat(),
                "ecoli": worst,  # actually enterococos NMP/100mL
                "url": url,
                "confidence": "oficial",
            }
            break
    return out


def _inea_parse_wide(wb, url):
    """Parse 'wide' sheets where dates are columns and rows are stations."""
    out = {}
    for bid, (sheet, code_prefixes) in INEA_WIDE.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        # Find header row: row whose col D... contains datetimes and col C
        # would equal "Estações" / blank. Use the row that has the most datetimes.
        header_row_idx = None
        best_count = 0
        for i, r in enumerate(rows[:10]):
            n = sum(1 for v in r if hasattr(v, "year"))
            if n > best_count:
                best_count = n
                header_row_idx = i
        if header_row_idx is None or best_count == 0:
            continue
        header = rows[header_row_idx]
        # Find data rows: must have a station code in col C matching a prefix.
        candidate_rows = []
        for r in rows[header_row_idx + 1:]:
            code = r[2]
            if not isinstance(code, str):
                continue
            code_s = code.strip()
            if any(code_s == p or code_s.startswith(p) for p in code_prefixes):
                candidate_rows.append(r)
        if not candidate_rows:
            continue
        # Find latest date column with a numeric value in any candidate row.
        n_cols = max(len(header), max(len(r) for r in candidate_rows))
        latest_idx = None
        latest_date = None
        latest_worst = None
        for ci in range(n_cols - 1, 2, -1):
            d = header[ci] if ci < len(header) else None
            if not hasattr(d, "year"):
                continue
            vals = []
            for r in candidate_rows:
                if ci >= len(r):
                    continue
                v = r[ci]
                if isinstance(v, (int, float)):
                    vals.append(int(v))
            if not vals:
                continue
            latest_idx = ci
            latest_date = d
            latest_worst = max(vals)
            break
        if latest_idx is None:
            continue
        status = _ent_status(latest_worst)
        if not status:
            continue
        out[bid] = {
            "status": status,
            "source": "INEA",
            "report_date": latest_date.date().isoformat(),
            "ecoli": latest_worst,
            "url": url,
            "confidence": "oficial",
        }
    return out


def fetch_inea():
    """Parse INEA's historical XLSX (Praias Monitoradas pelo INEA desde 2005)
    instead of the weekly image-only PDF. The XLSX is linked from the
    balneabilidade landing page and updated monthly with the latest sampling
    rounds. We pick the most recent date with a numeric reading per beach
    and classify using CONAMA 274/2000 (Enterococos ≤100 NMP/100mL = própria).
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("  INEA: openpyxl not installed — skipping", file=sys.stderr)
        return {}

    try:
        html = curl(INEA_LANDING)
    except subprocess.CalledProcessError as e:
        print(f"  INEA: landing err {e}", file=sys.stderr)
        return {}
    xlsx_url = _inea_find_xlsx_url(html)
    if not xlsx_url:
        print("  INEA: no XLSX link on landing page", file=sys.stderr)
        return {}
    if xlsx_url.startswith("/"):
        xlsx_url = "https://www.inea.rj.gov.br" + xlsx_url

    try:
        blob = curl(xlsx_url, binary=True)
    except subprocess.CalledProcessError as e:
        print(f"  INEA: download err {e}", file=sys.stderr)
        return {}
    if not blob[:2] == b"PK":  # xlsx = zip
        print("  INEA: not a zip/xlsx", file=sys.stderr)
        return {}

    import io
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True,
                                    data_only=True)
    except Exception as e:
        print(f"  INEA: openpyxl err {e}", file=sys.stderr)
        return {}

    print(f"  INEA: parsing {xlsx_url}", file=sys.stderr)
    out = {}
    try:
        out.update(_inea_parse_zsul(wb, xlsx_url))
    except Exception as e:
        print(f"  INEA: zsul parse err {e}", file=sys.stderr)
    try:
        out.update(_inea_parse_wide(wb, xlsx_url))
    except Exception as e:
        print(f"  INEA: wide parse err {e}", file=sys.stderr)
    return out


# -----------------------------------------------------------------------------
# INEMA (Bahia)
# -----------------------------------------------------------------------------
# INEMA's PDFs are at unpredictable filenames on homologa.ba.gov.br
# (which is also often unreachable from outside BA). However, every Friday
# INEMA publishes a *news article* with the full bulletin in prose form,
# listing all sample points by status. We discover the latest such article
# via the on-site search and parse the body text.

INEMA_SEARCH = "https://www.ba.gov.br/inema/busca-solr-multisite?keys=balneabilidade"
INEMA_BASE = "https://www.ba.gov.br"

# beach_id → list of (substring, optional context hint) — substring must hit
# one of the próprios / impróprios lists in the article.
INEMA_POINTS = {
    "morro-terceira": ["terceira praia", "3ª praia", "3a praia"],
    "salvador-itapua": ["itapuã", "itapua"],
    "salvador-stella-maris": ["stella maris"],
    "porto-seguro-taperapua": ["taperapuã", "taperapua"],
}

# Try direct PDF URL patterns as a secondary strategy (in case INEMA
# ever links them from a future news article). Strategy:
#   1) discover candidate PDF urls from any recent news article body
#   2) fall back to guessing the most recent Fridays of current/last month
INEMA_PDF_MUNI = {
    "morro-terceira": "Cairu",
    "salvador-itapua": "Salvador",
    "salvador-stella-maris": "Salvador",
    "porto-seguro-taperapua": "Porto Seguro",
}


def _inema_find_latest_news():
    """Return list of (url, slug-date-hint) for the most recent
    balneabilidade news articles, newest first."""
    try:
        html_text = curl(INEMA_SEARCH)
    except subprocess.CalledProcessError as e:
        print(f"  INEMA: search err {e}", file=sys.stderr)
        return []
    links = re.findall(
        r'href="(https?://[^"]*inema/noticias/[^"]*balneabilidade[^"]*)"',
        html_text, re.I,
    )
    # Dedup, keep order, prefer ones whose slug describes a "divulga situação"
    # / "praias" listing (vs. methodology explainers).
    seen = set()
    ordered = []
    for u in links:
        u = u.replace("&amp;", "&")
        if u in seen:
            continue
        seen.add(u)
        ordered.append(u)
    # Prefer divulga/situacao slugs first; then anything with a YYYY-MM path.
    def score(u):
        s = 0
        if "divulga" in u or "situacao" in u or "praias" in u:
            s += 10
        # newer YYYY-MM path scores higher (lexicographic on the matched slice)
        m = re.search(r"/(\d{4}-\d{2})/", u)
        if m:
            s += 1
        return (s, u)
    ordered.sort(key=score, reverse=True)
    return ordered


def _inema_strip(html_text):
    """Reduce news HTML to a single whitespace-collapsed plain-text blob."""
    h = re.sub(r"<(script|style)[^>]*>.*?</\1>", "", html_text,
               flags=re.S | re.I)
    h = re.sub(r"<[^>]+>", " ", h)
    # Decode HTML entities the cheap way.
    from html import unescape
    h = unescape(h)
    h = re.sub(r"\s+", " ", h).strip()
    return h


def _inema_parse_date(text):
    m = re.search(r"(\d{2}/\d{2}/\d{4})", text)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%d/%m/%Y").date().isoformat()
    except ValueError:
        return None


def _inema_classify_point(text_low, needles):
    """Look for any needle in text_low and decide própria/imprópria based on
    which list segment it falls in. INEMA articles always interleave
    'próprio(s/as)' and 'impróprio(s/as)' paragraph markers — we find the
    nearest preceding marker before the needle hit.

    For the Morro de São Paulo complex, the article uses a summary clause
    instead of enumerating each praia ("complexo de Morro de São Paulo
    apresentou condição própria — à exceção da 1ª praia"). We special-case
    Terceira Praia below.
    """
    for needle in needles:
        idx = text_low.find(needle)
        if idx < 0:
            continue
        window = text_low[:idx]
        last_propr = max(window.rfind("próprio"), window.rfind("proprio"),
                         window.rfind("próprios"), window.rfind("próprias"),
                         window.rfind("condição própria"),
                         window.rfind("condicao propria"))
        last_impr = max(window.rfind("impróprio"), window.rfind("improprio"),
                        window.rfind("impróprios"), window.rfind("impróprias"),
                        window.rfind("condição imprópria"))
        if last_propr < 0 and last_impr < 0:
            return None
        return "propria" if last_propr > last_impr else "impropria"
    return None


def _inema_parse_news(text, url):
    """Return {beach_id: entry} parsed from a balneabilidade news article."""
    report_date = _inema_parse_date(text)
    low = text.lower()
    out = {}

    for bid, needles in INEMA_POINTS.items():
        if bid == "morro-terceira":
            # Special case: the article describes the Morro de São Paulo
            # complex as a whole. Terceira Praia is própria unless the
            # article explicitly excludes "3ª praia" / "terceira".
            if "morro de são paulo" not in low and "morro de sao paulo" not in low:
                continue
            # Default to própria; flip if Terceira is mentioned as exception.
            status = "propria"
            # Look for explicit exceptions naming Terceira / 3ª.
            for kw in ("terceira praia", "3ª praia", "3a praia"):
                k = low.find(kw)
                if k < 0:
                    continue
                # Check the nearest preceding "exceção" or "imprópria" within
                # ~120 chars.
                ctx = low[max(0, k - 200):k + 60]
                if ("exceção" in ctx or "excecao" in ctx
                        or "imprópri" in ctx or "improp" in ctx):
                    status = "impropria"
                    break
            out[bid] = {
                "status": status,
                "source": "INEMA",
                "report_date": report_date,
                "ecoli": None,
                "url": url,
                "confidence": "oficial",
            }
            continue

        status = _inema_classify_point(low, needles)
        if not status:
            continue
        out[bid] = {
            "status": status,
            "source": "INEMA",
            "report_date": report_date,
            "ecoli": None,
            "url": url,
            "confidence": "oficial",
        }

    return out


def _inema_try_pdfs(news_text):
    """If the news article happens to link any boletim PDFs, try to parse
    those for E. coli numbers. Best-effort: returns {beach_id: ecoli_int}."""
    pdf_urls = re.findall(
        r'(https?://[^\s"<>]+Boletim[^\s"<>]*\.pdf)', news_text, re.I)
    ecoli = {}
    for url in pdf_urls[:6]:
        try:
            pdf = curl(url, binary=True)
        except subprocess.CalledProcessError:
            continue
        if not pdf.startswith(b"%PDF"):
            continue
        try:
            text = pdf_to_text(pdf)
        except Exception:
            continue
        low = text.lower()
        for bid, needles in INEMA_POINTS.items():
            if bid in ecoli:
                continue
            for needle in needles:
                idx = low.find(needle)
                if idx < 0:
                    continue
                # Grab line containing the needle.
                line_start = low.rfind("\n", 0, idx) + 1
                line_end = low.find("\n", idx)
                if line_end < 0:
                    line_end = len(low)
                line = text[line_start:line_end]
                m = re.search(r"\b(\d{1,7})\b\s*(?:nmp|/100|colônias|colonias)?",
                              line, re.I)
                if m:
                    try:
                        ecoli[bid] = int(m.group(1))
                    except ValueError:
                        pass
                break
    return ecoli


def fetch_inema():
    news_urls = _inema_find_latest_news()
    if not news_urls:
        print("  INEMA: no news articles found", file=sys.stderr)
        return {}
    last_err = None
    for nu in news_urls[:4]:
        try:
            raw = curl(nu)
        except subprocess.CalledProcessError as e:
            last_err = e
            continue
        text = _inema_strip(raw)
        out = _inema_parse_news(text, nu)
        if not out:
            continue
        print(f"  INEMA: parsing {nu}", file=sys.stderr)
        # Best-effort: enrich with E. coli numbers from any linked PDFs.
        try:
            ecoli = _inema_try_pdfs(raw)
            for bid, val in ecoli.items():
                if bid in out:
                    out[bid]["ecoli"] = val
        except Exception as e:
            print(f"  INEMA: PDF enrich skipped: {e}", file=sys.stderr)
        return out
    if last_err:
        print(f"  INEMA: last error {last_err}", file=sys.stderr)
    return {}


# -----------------------------------------------------------------------------
# CPRH (Pernambuco)
# -----------------------------------------------------------------------------
# Weekly PDFs listed at /informativos-semanais-YYYY/. We hit the parent index
# page (which links to the current-year page), then grab the newest PDF and
# substring-match beach names against the table rows.

CPRH_INDEX = ("https://www2.cprh.pe.gov.br/monitoramento-ambiental/"
              "balneabilidade/")

# beach_id → list of substrings that uniquely identify the beach row.
CPRH_POINTS = {
    "recife-boa-viagem": ["boa viagem"],
    "ipojuca-porto-galinhas": ["porto de galinhas"],
}


def _cprh_latest_pdf_url():
    # First try the year-specific index pages linked from the parent.
    try:
        landing = curl(CPRH_INDEX)
    except subprocess.CalledProcessError as e:
        print(f"  CPRH: landing err {e}", file=sys.stderr)
        return None
    year_pages = re.findall(
        r'href="(https?://[^"]*informativos?-semanais?-\d{4}/?)"',
        landing, re.I)
    pdfs = []
    pages_to_scan = list(dict.fromkeys(year_pages)) or [CPRH_INDEX]
    for page in pages_to_scan[:3]:
        try:
            html = curl(page)
        except subprocess.CalledProcessError:
            continue
        found = re.findall(
            r'href="(https?://[^"]*informativo-balneabilidade-\d+_\d{4}[^"]*\.pdf)"',
            html, re.I)
        pdfs.extend(found)
        if pdfs:
            break
    if not pdfs:
        # Fallback: any PDF on landing.
        pdfs = re.findall(r'href="([^"]+\.pdf)"', landing, re.I)
    if not pdfs:
        return None

    def week_num(u):
        m = re.search(r"informativo-balneabilidade-(\d+)_(\d{4})", u, re.I)
        if not m:
            return (0, 0)
        return (int(m.group(2)), int(m.group(1)))

    pdfs.sort(key=week_num, reverse=True)
    return pdfs[0]


def fetch_cprh():
    url = _cprh_latest_pdf_url()
    if not url:
        print("  CPRH: no PDF found", file=sys.stderr)
        return {}
    try:
        pdf = curl(url, binary=True)
    except subprocess.CalledProcessError as e:
        print(f"  CPRH: download err {e}", file=sys.stderr)
        return {}
    if not pdf.startswith(b"%PDF"):
        return {}
    try:
        text = pdf_to_text(pdf)
    except Exception as e:
        print(f"  CPRH: pdftotext failed: {e}", file=sys.stderr)
        return {}

    # Date: look for a "DATA DA COLETA: dd/mm/yyyy" or any dd/mm/yyyy.
    report_date = None
    m = re.search(r"COLETA[^0-9]*(\d{2}/\d{2}/\d{4})", text, re.I)
    if not m:
        m = re.search(r"(\d{2}/\d{2}/\d{4})", text)
    if m:
        try:
            report_date = datetime.strptime(
                m.group(1), "%d/%m/%Y").date().isoformat()
        except ValueError:
            pass

    print(f"  CPRH: parsing {url}", file=sys.stderr)
    out = {}
    for line in text.splitlines():
        low = line.lower()
        status = None
        if "impróp" in low or "improp" in low:
            status = "impropria"
        elif "própr" in low or "propr" in low:
            status = "propria"
        if not status:
            continue
        for bid, needles in CPRH_POINTS.items():
            for needle in needles:
                if needle in low:
                    # Prefer keeping the first hit per beach (matches the
                    # primary sample point listed).
                    if bid not in out:
                        out[bid] = {
                            "status": status,
                            "source": "CPRH",
                            "report_date": report_date,
                            "ecoli": None,
                            "url": url,
                            "confidence": "oficial",
                        }
                    break
    return out


# -----------------------------------------------------------------------------
# SEMACE (Ceará)
# -----------------------------------------------------------------------------
# No JSON API was found on the boletim page — only a list of weekly PDF links
# under wp-content/uploads/.../Boletim-*.pdf. The wp-json endpoint exists but
# only returns the page wrapper, not bulletin data. So we parse the PDF.

SEMACE_LANDING = "https://www.semace.ce.gov.br/boletim-de-balneabilidade/"

SEMACE_POINTS = {
    "fortaleza-iracema-meireles": ["iracema", "meireles"],
}

_STATUS_PRIORITY = {"impropria": 2, "alerta": 1, "propria": 0}


def fetch_semace():
    try:
        html = curl(SEMACE_LANDING)
    except subprocess.CalledProcessError as e:
        print(f"  SEMACE: landing err {e}", file=sys.stderr)
        return {}
    pdfs = re.findall(
        r'href="(https?://[^"]*?Boletim-\d+[^"]*\.pdf)"', html, re.I)
    if not pdfs:
        pdfs = re.findall(r'href="([^"]+\.pdf)"', html, re.I)
    if not pdfs:
        print("  SEMACE: no PDF links", file=sys.stderr)
        return {}

    def boletim_date(u):
        m = re.search(r"Boletim-(\d{8})", u)
        return m.group(1) if m else ""

    pdfs.sort(key=boletim_date, reverse=True)
    for url in pdfs[:4]:
        try:
            pdf = curl(url, binary=True)
        except subprocess.CalledProcessError:
            continue
        if not pdf.startswith(b"%PDF"):
            continue
        try:
            text = pdf_to_text(pdf)
        except Exception as e:
            print(f"  SEMACE: pdftotext err {e}", file=sys.stderr)
            continue
        print(f"  SEMACE: parsing {url}", file=sys.stderr)

        report_date = None
        m = re.search(r"Boletim-(\d{4})(\d{2})(\d{2})", url)
        if m:
            report_date = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        else:
            m2 = re.search(r"(\d{2}/\d{2}/\d{4})", text)
            if m2:
                try:
                    report_date = datetime.strptime(
                        m2.group(1), "%d/%m/%Y").date().isoformat()
                except ValueError:
                    pass

        out = {}
        # SEMACE format: each line ends with a trailing " P" or " I"
        # (própria / imprópria) as classification token.
        for line in text.splitlines():
            low = line.lower()
            hit_bid = None
            for bid, needles in SEMACE_POINTS.items():
                for needle in needles:
                    if needle in low:
                        hit_bid = bid
                        break
                if hit_bid:
                    break
            if not hit_bid:
                continue
            status = None
            stripped = line.rstrip()
            # Last single-letter token after whitespace.
            mt = re.search(r"\s([PIp i])\s*$", stripped)
            if mt:
                tok = mt.group(1).upper()
                if tok == "P":
                    status = "propria"
                elif tok == "I":
                    status = "impropria"
            if not status:
                # Fallback: look for the words in the line.
                if "impróp" in low or "improp" in low:
                    status = "impropria"
                elif "própr" in low or "propr" in low:
                    status = "propria"
            if not status:
                continue
            prev = out.get(hit_bid)
            if (prev is None or
                    _STATUS_PRIORITY[status] > _STATUS_PRIORITY[prev["status"]]):
                out[hit_bid] = {
                    "status": status,
                    "source": "SEMACE",
                    "report_date": report_date,
                    "ecoli": None,
                    "url": url,
                    "confidence": "oficial",
                }
        if out:
            return out
    return {}


# -----------------------------------------------------------------------------
# IMA-AL (Alagoas)
# -----------------------------------------------------------------------------
# The relatórios page loads documents asynchronously via wp-admin/admin-ajax.php
# with category 'balneabilidade-das-praias'. We fetch the AJAX response,
# extract the newest PDF link, and parse it.

IMA_AL_LANDING = ("https://www2.ima.al.gov.br/laboratorio/"
                  "relatorios-de-balneabilidade/balneabilidade-de-praias/")
IMA_AL_AJAX = "https://www2.ima.al.gov.br/wp-admin/admin-ajax.php"

IMA_AL_POINTS = {
    "maceio-pajucara": ["pajuçara", "pajucara", "ponta verde"],
}


def _ima_al_latest_pdf():
    # We need a fresh nonce: grab from the landing page.
    try:
        page = curl(IMA_AL_LANDING)
    except subprocess.CalledProcessError as e:
        print(f"  IMA-AL: landing err {e}", file=sys.stderr)
        return None
    m = re.search(r'documentos_vars\s*=\s*\{[^}]*"nonce"\s*:\s*"([a-f0-9]+)"',
                  page)
    nonce = m.group(1) if m else ""
    # POST to admin-ajax
    data = (f"action=carregar_documentos&nonce={nonce}"
            f"&categoria=balneabilidade-das-praias&per_page=5"
            f"&orderby=date&page=1")
    try:
        out = subprocess.run(
            ["curl", "-sSfL", "-A", UA, "-m", str(TIMEOUT),
             "-X", "POST", "--data", data, IMA_AL_AJAX],
            capture_output=True, check=True, timeout=TIMEOUT + 5,
        ).stdout.decode("utf-8", errors="replace")
    except subprocess.CalledProcessError as e:
        print(f"  IMA-AL: ajax err {e}", file=sys.stderr)
        return None
    pdfs = re.findall(r'href="(https?://[^"]+\.pdf)"', out, re.I)
    if not pdfs:
        return None
    return pdfs[0]


def fetch_ima_al():
    url = _ima_al_latest_pdf()
    if not url:
        print("  IMA-AL: no PDF found", file=sys.stderr)
        return {}
    try:
        pdf = curl(url, binary=True)
    except subprocess.CalledProcessError as e:
        print(f"  IMA-AL: download err {e}", file=sys.stderr)
        return {}
    if not pdf.startswith(b"%PDF"):
        return {}
    try:
        text = pdf_to_text(pdf)
    except Exception as e:
        print(f"  IMA-AL: pdftotext err {e}", file=sys.stderr)
        return {}

    print(f"  IMA-AL: parsing {url}", file=sys.stderr)
    report_date = None
    m = re.search(r"REAB-(\d+)-(\d{4})", url)
    # Prefer date in body if present.
    md = re.search(r"(\d{2}/\d{2}/\d{4})", text)
    if md:
        try:
            report_date = datetime.strptime(
                md.group(1), "%d/%m/%Y").date().isoformat()
        except ValueError:
            pass

    out = {}
    for line in text.splitlines():
        low = line.lower()
        status = None
        if "impróp" in low or "improp" in low:
            status = "impropria"
        elif "própr" in low or "propr" in low:
            status = "propria"
        if not status:
            continue
        for bid, needles in IMA_AL_POINTS.items():
            if any(n in low for n in needles):
                prev = out.get(bid)
                if (prev is None or
                        _STATUS_PRIORITY[status] >
                        _STATUS_PRIORITY[prev["status"]]):
                    out[bid] = {
                        "status": status,
                        "source": "IMA-AL",
                        "report_date": report_date,
                        "ecoli": None,
                        "url": url,
                        "confidence": "oficial",
                    }
                break
    return out


# -----------------------------------------------------------------------------
# FEPAM (Rio Grande do Sul)
# -----------------------------------------------------------------------------
# FEPAM's Projeto Balneabilidade is *seasonal* — weekly Friday bulletins are
# only published during "Operação Verão Total" (roughly Dec → late Feb).
# Outside that window the dedicated app host (balneabilidade.rs.gov.br) is
# empty and no boletim is produced, so we short-circuit with an off_season
# stub instead of pointlessly hammering the site.

FEPAM_SEARCH = "https://www.fepam.rs.gov.br/?s=balneabilidade"
FEPAM_HOST = "https://www.fepam.rs.gov.br"

FEPAM_POINTS = {
    "tramandai": ["tramandaí", "tramandai"],
    "capao-da-canoa": ["capão da canoa", "capao da canoa"],
}


def _fepam_in_season(today=None):
    today = today or date.today()
    # Dec, Jan, Feb only. Mar–Nov is off-season.
    return today.month in (12, 1, 2)


def _fepam_off_season_stub():
    out = {}
    today = date.today().isoformat()
    for bid in FEPAM_POINTS:
        out[bid] = {
            "status": "sem_dados",
            "source": "FEPAM",
            "report_date": None,
            "ecoli": None,
            "url": FEPAM_SEARCH,
            "confidence": "estimativa",
            "off_season": True,
            "note": "FEPAM publica boletins apenas durante a Operação Verão Total (dez–fev).",
            "checked_at": today,
        }
    return out


def _fepam_classify_window(text_low, idx):
    """Find the nearest preceding propria/impropria marker before idx."""
    window = text_low[:idx]
    last_p = max(window.rfind("própri"), window.rfind("propri"))
    last_i = max(window.rfind("impróp"), window.rfind("improp"))
    if last_p < 0 and last_i < 0:
        return None
    return "impropria" if last_i > last_p else "propria"


def fetch_fepam():
    if not _fepam_in_season():
        print("  FEPAM: off-season (May 19) → sem_dados stubs", file=sys.stderr)
        return _fepam_off_season_stub()

    # In-season: try to discover the most-recent boletim news article via the
    # WP search page and parse the prose listing of próprios/impróprios pts.
    try:
        html_text = curl(FEPAM_SEARCH)
    except subprocess.CalledProcessError as e:
        print(f"  FEPAM: search err {e}", file=sys.stderr)
        return _fepam_off_season_stub()

    links = re.findall(
        r'href="(https?://[^"]*fepam\.rs\.gov\.br/[^"]*balneabilidade[^"]*)"',
        html_text, re.I)
    rel = re.findall(r'href="(/[^"]*balneabilidade[^"]*)"', html_text, re.I)
    links += [FEPAM_HOST + r for r in rel]
    # Dedupe, prefer slugs that look like boletim/projeto/litoral.
    seen, ordered = set(), []
    for u in links:
        u = u.replace("&amp;", "&")
        if u in seen or u.endswith(".css") or u.endswith(".js"):
            continue
        seen.add(u)
        ordered.append(u)
    ordered.sort(
        key=lambda u: ("boletim" in u or "projeto" in u or "litoral" in u),
        reverse=True,
    )

    for nu in ordered[:5]:
        try:
            raw = curl(nu)
        except subprocess.CalledProcessError:
            continue
        # Strip HTML cheaply.
        h = re.sub(r"<(script|style)[^>]*>.*?</\1>", "", raw,
                   flags=re.S | re.I)
        h = re.sub(r"<[^>]+>", " ", h)
        from html import unescape
        text = unescape(re.sub(r"\s+", " ", h)).strip()
        low = text.lower()
        if "balneabilidade" not in low:
            continue
        report_date = None
        m = re.search(r"(\d{2}/\d{2}/\d{4})", text)
        if m:
            try:
                report_date = datetime.strptime(
                    m.group(1), "%d/%m/%Y").date().isoformat()
            except ValueError:
                pass

        out = {}
        for bid, needles in FEPAM_POINTS.items():
            for n in needles:
                idx = low.find(n)
                if idx < 0:
                    continue
                status = _fepam_classify_window(low, idx)
                if not status:
                    continue
                out[bid] = {
                    "status": status,
                    "source": "FEPAM",
                    "report_date": report_date,
                    "ecoli": None,
                    "url": nu,
                    "confidence": "oficial",
                }
                break
        if out:
            print(f"  FEPAM: parsing {nu}", file=sys.stderr)
            # Fill in misses with off-season-style stubs but keep season flag off.
            for bid in FEPAM_POINTS:
                out.setdefault(bid, {
                    "status": "sem_dados",
                    "source": "FEPAM",
                    "report_date": report_date,
                    "ecoli": None,
                    "url": nu,
                    "confidence": "estimativa",
                })
            return out

    print("  FEPAM: in-season but no boletim parsed", file=sys.stderr)
    return _fepam_off_season_stub()


# -----------------------------------------------------------------------------
# Build
# -----------------------------------------------------------------------------

def build():
    by_beach = {}

    for name, fn in [("IMA-SC", fetch_ima_sc),
                     ("CETESB", fetch_cetesb),
                     ("INEA",   fetch_inea),
                     ("INEMA",  fetch_inema),
                     ("CPRH",   fetch_cprh),
                     ("SEMACE", fetch_semace),
                     ("IMA-AL", fetch_ima_al),
                     ("FEPAM",  fetch_fepam)]:
        try:
            res = fn() or {}
            print(f"  {name}: {len(res)} beaches resolved", file=sys.stderr)
            by_beach.update(res)
        except Exception as e:
            print(f"  {name}: ERR {e}", file=sys.stderr)

    # Fill sem_dados stubs for anything missing.
    for bid in ALL_BEACH_IDS:
        if bid in by_beach:
            continue
        by_beach[bid] = {
            "status": "sem_dados",
            "source": DEFAULT_SOURCE.get(bid, "—"),
            "report_date": None,
            "ecoli": None,
            "url": None,
            "confidence": "estimativa",
        }

    payload = {"fetched_at": int(time.time()), "by_beach": by_beach}
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    json.dump(payload, open(OUT, "w"), ensure_ascii=False, indent=2)

    real = sum(1 for v in by_beach.values() if v["status"] != "sem_dados")
    print(f"Wrote {OUT} — {real}/{len(by_beach)} with real data", file=sys.stderr)
    for bid, v in by_beach.items():
        print(f"  {bid:28} {v['status']:10} {v['source']}", file=sys.stderr)


if __name__ == "__main__":
    build()
